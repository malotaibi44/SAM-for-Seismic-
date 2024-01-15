# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 01:47:37 2024

@author: Mohammed
"""

import sys
from time import sleep
using_colab = False
import numpy as np
import torch
import matplotlib.pyplot as plt
import cv2
from openpyxl import Workbook
import statistics
import os
import random
from IPython import get_ipython
from matplotlib.backend_bases import MouseButton
import time
from PIL import Image 
from segment_anything import sam_model_registry, SamPredictor
import matplotlib
#%%
# names=os.listdir('C:/Users/Mohammed/Downloads/saltdome')
# labels=os.listdir('C:/Users/Mohammed/Downloads/labels')
def show_mask(mask, ax, random_color=False):
    if random_color:
        color = np.concatenate([np.random.random(3), np.array([0.6])], axis=0)
    else:
        color = np.array([30/255, 144/255, 255/255, 0.6])
    h, w = mask.shape[-2:]
    mask_image = mask.reshape(h, w, 1) * color.reshape(1, 1, -1)
    ax.imshow(mask_image)
    
def show_points(coords, labels, ax, marker_size=50):
    
    pos_points = coords[labels==1]
    neg_points = coords[labels==0]
    ax.scatter(pos_points[:, 0], pos_points[:, 1], color='green', marker='*', s=marker_size, edgecolor='white', linewidth=1.25)
    ax.scatter(neg_points[:, 0], neg_points[:, 1], color='red', marker='*', s=marker_size, edgecolor='white', linewidth=1.25)   
      

def closetn(node, nodes):
    nodes = np.asarray(nodes)
    deltas = nodes - node
    dist_2 = np.einsum('ij,ij->i', deltas, deltas)
    return np.argmin(dist_2)



sys.path.append("..")


sam_checkpoint = 'sam_vit_h_4b8939.pth'
model_type = "vit_h"

device = "cuda" if torch.cuda.is_available() else 'cpu'

sam = sam_model_registry[model_type](checkpoint=sam_checkpoint)
sam.to(device=device)

predictor = SamPredictor(sam)

names=np.load("samples_.npy")
labels=np.load("lbls_.npy")

#%%

first=input("Do you want to load previous work? -y -n\n")
while first!='n' and first!='y':
    first=input("Chose y or n, Do you want to load previous work? -y -n\n")
if first=='n':
    wb = Workbook()
    ws = wb.active
    ws['A1']='slice'
    ws['B1']='# green dots of best'
    ws['C1']='# red dots of best '
    ws['D1']='SD of green of best '
    ws['E1']='SD of red of best'
    ws['F1']='best score'
    # for i in range(10):
    #     ws[i+'1']='# green dots of '+str(i)
    #     ws[chr(72+i*5)+'1']='# red dots of '+str(i)
    #     ws[chr(73+i*5)+'1']='SD of green of '+str(i)
    #     ws[chr(74+i*5)+'1']='SD of red of '+str(i)
    #     ws[chr(75+i*5)+'1']='score of '+str(i)
    for i in range(9):
        coun=1
        for col in ws.iter_cols(min_row=1,max_row=1, max_col=12+i*5, min_col=7+i*5):
            if coun==1:
                ws[col[0].coordinate]='# green dots of '+str(i+2)
            elif coun==2:
                ws[col[0].coordinate]='# red dots of '+str(i+2)
            elif coun==3:
                ws[col[0].coordinate]='SD of X of '+str(i+2)
            elif coun==4:
                ws[col[0].coordinate]='SD of Y of '+str(i+2)
            elif coun==5:
                ws[col[0].coordinate]='score of '+str(i+2)
            coun+=1
    name = input("Type your name:\n")

    if not os.path.exists(name):
        os.makedirs(name)
        os.makedirs(os.path.join(name,"masks"))
        os.makedirs(os.path.join(name,"points"))
        os.makedirs(os.path.join(name,"sorts"))
        
    c=0
    tim=0
    t=time.time()
    
else:
    from openpyxl import load_workbook
    name=input("what is your name?\n")
    wb=load_workbook(os.path.join(name,name+".xlsx"))
    ws=wb.active
    c=len(os.listdir(os.path.join(name,"masks")))
    f=open(os.path.join(name,"time.txt"),'r')
    
    tim=f.readline()
    t=time.time()
    f.close()







 #### change that later 
print(c)
f=False
## start looping through samples: 
while c<150 and not f:
    msk=[] # masks for each samples 

    gp=[] #green points
    rp=[] # red points 
    image = names[c] # samples c
    ws['A'+str(c+2)]=str(c) # samples name on excel 
    image=cv2.cvtColor((np.array(((image+1)/2)*255,dtype='uint8')), cv2.COLOR_GRAY2RGB)
    label=labels[c] # GT for sample c 
    rmv=False
    mask=0
    # image=np.array(((image+1)/2)*255,dtype='uint8') 
    predictor.set_image(image)
    inc=""
    co=0
    bs=0
    score=[]
  
    stdx=[]
    stdy=[]
    ng=[]
    nr=[]        
    green=[]
    red=[]
    greenx = []

    redx = []
    greeny = []
    redy = []
    # label=plt.imread('C:/Users/Mohammed/Downloads/labels/'+labels[c])
    label=label==5

    #matplotlib.use('TkAgg')
    try:
        
        matplotlib.use('Qt5Agg')
    except:
        matplotlib.use('TkAgg')

    while inc!="y":
        s=0 # this is for the score 
        count=1 # to count the score max 
        lessfive=0
        current_color = 'green'
        #get_ipython().run_line_magic('matplotlib', 'qt')
        fig, ax = plt.subplots(1,3,figsize=(15,7))
        if green and red :
            ax[0].plot(greenx, greeny, 'go', markersize=5)
            ax[1].plot(greenx, greeny, 'go', markersize=5)
            ax[0].plot(redx, redy, 'ro', markersize=5)
            ax[1].plot(redx, redy, 'ro', markersize=5)
            plt.draw()
        def onclose(event):
            fig.canvas.stop_event_loop()
            fig.canvas.mpl_disconnect(cid)
        def onclick(event):
            global count
            global green
            global red
            global greenx
            global redx
            global greeny
            global redy
            global label
            global mask
            global lessfive
            if event.xdata is not None and event.ydata is not None:


                x, y = int(event.xdata), int(event.ydata)
                print(not x)
                print(not y)
                # if not x or not y:
                #     inc=input("do you wish to continue?")
                #     f=True

                if event.button is MouseButton.LEFT :
                    if current_color == 'green':
                      
                        green.append((x,y))
                        greenx.append(x)
                        
                        
                        greeny.append(y)
                        ax[0].plot(x, y, 'go', markersize=5)
                        ax[1].plot(x, y, 'go', markersize=5)
                        plt.draw()

                    else:
                        red.append((x,y))
                        redx.append(x)
                        
                        redy.append(y)
                        ax[0].plot(x, y, 'ro', markersize=5)
                        ax[1].plot(x, y, 'ro', markersize=5)
                        plt.draw()
                    
                elif event.button is MouseButton.RIGHT:

                    if not green and not red: 
                        print("no points to delete")
                    elif green:
                        print(current_color)
                        if current_color == 'green':
                            #print("g",len(green))
                            
                            indx=closetn((x,y),green)
                            print(indx)
                            for line in ax[0].lines:
                                if len(line.get_xdata())>0:
                                    if line.get_xdata()[0]==green[indx][0] and line.get_ydata()[0]==green[indx][1]:
                                        
                                        #print("Here1")
                                        line.set_data([],[])
                                        break
                            for line in ax[1].lines:
                                if len(line.get_xdata())>0:
                                    if line.get_xdata()[0]==green[indx][0] and line.get_ydata()[0]==green[indx][1]:
                                        #print("Here2")
                                        line.set_data([],[])
                                        break
                            del green[indx]
                            del greenx[indx]
                            
                            del greeny[indx]

                            #ax[0].plot(x, y, 'go', markersize=5)
                            #ax[1].plot(x, y, 'go', markersize=5)

                            plt.draw()
                    elif red:
                        print(current_color)
                        indx=closetn((x,y),red)
                        
                        for line in ax[0].lines:
                            if len(line.get_xdata())>0:
                                print()
                                if line.get_xdata()[0]==red[indx][0] and line.get_ydata()[0]==red[indx][1]:
                                    line.set_data([],[])
                                    break
                        for line in ax[1].lines:
                            if len(line.get_xdata())>0:
                                if line.get_xdata()[0]==red[indx][0] and line.get_ydata()[0]==red[indx][1]:
                                    line.set_data([],[])
                                    break
                        #ax[0].plot(x, y, 'ro', markersize=5)
                        #ax[1].plot(x, y, 'ro', markersize=5)                            
                        #ax[0].set_offsets(red)
                        #a.set_offsets(red)
                        del red[indx]
                        del redx[indx]
                        
                        del redy[indx]
                        plt.draw()

                if green and red:
                    global s
                    print("green:",green)
                    print("red:",red)
                    
                    
                    input_point=np.concatenate((green,red))
                    input_label=np.concatenate(([1]*len(green),[0]*len(red)))
                   
                    masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                    )
                    
                    mask=masks[0]
                    
                    #get_ipython().run_line_magic('matplotlib', 'inline')
                    ax[2].clear()
                    ax[2].imshow(image)
                    show_mask(mask, ax[2])
                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s=0
                    else:
                      s=intersection / union
                    #ws[chr(68)+str(c+2)]=str(bs) # start at cell D(c)
                    show_points(input_point, input_label, ax[2])
                    msg=""
               
                    if len(score)==0:
                        maxx=0
                    else:
                        maxx=max(score)
                                        
                    score.append(s)
                    gp.append(np.multiply(green,1))
                    
                    rp.append(np.multiply(red,1))
                    ng.append(len(greenx))
                    nr.append(len(redx))
                    stdx.append(statistics.pstdev(np.concatenate((greenx,redx))))
                    stdy.append(statistics.pstdev(np.concatenate((greeny,redy))))
                    print("up count",count)
                    if maxx>=s:
                        print("\nheere\n")
                        if count>=5:
                            print("lesssss")
                            #msg="\nNo better score is achieved in the last 5 attempts. Start round 2 from scra\nThe best score ("+str(round(max(score),2))+") is saved"
                            lessfive+=1
                        else:
                            print("less")
                            count+=1
                    elif  maxx<s:
                        print("more")
                        count=1
                    if lessfive==1:
                        print("Heeeeeeereless")
                        msg=" (round 2) "
                    plt.title(f"Score: {(intersection / union):.3f}"+msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask,5))
                    print("lessfive",lessfive)
                    print("scores:",score)
                    if lessfive==1:
                        lessfive+=1
                        for line in ax[0].lines:
                            line.set_data([], [])
                        for line in ax[1].lines:
                            line.set_data([], [])
                        green = []
                        red = []
                        greenx = []
                        redx = []
                        greeny = []
                        redy = []
                        plt.draw()
                        ax[2].clear()
                        ax[2].imshow(image)
                        show_mask(mask, ax[2])
                        count = 1
                        print("below count", count)
                        plt.title("No better score is achieved in the last 5 attempts. Start round 2 from scratch")
                    elif lessfive==3:
                        print("The window closed because you did not achieve a better score after 5 consecutive clicks in the 2nd round")
                        plt.close()
        # Create a function to toggle between green and red dots
        def toggle_color(event):
            global green
            global red
            global greenx
            global redx
            global greeny
            global redy
            global current_color
            global count
            if event.key == 'g':
                current_color = 'green'
                print("Switched to GREEN dot mode.")
                
            elif event.key == 'r':
                current_color = 'red'
                print("Switched to RED dot mode.")
            elif event.key == ' ':
                for line in ax[0].lines:
                    line.set_data([],[])
                for line in ax[1].lines:
                    line.set_data([],[])
                green=[]
                red=[]
                greenx=[]
                redx=[]
                greeny=[]
                redy=[]
                plt.draw()
                ax[2].clear()
                ax[2].imshow(image)
                show_mask(mask, ax[2])
                count=1
                print("below count",count)
                plt.title("No better score is achieved in the last 5 attempts. Start round 2 from scratch")
    # Create a figure and display the image
    
    
        a=ax[0].plot()
        b=ax[1].plot()
        ax[0].imshow(image)
        ax[1].imshow(label)
        # Connect mouse click and keyboard key events
        fig.canvas.mpl_connect('button_press_event', onclick)
        #fig.canvas.start_event_loop(timeout=-5)
        fig.canvas.mpl_connect('key_press_event', toggle_color)
        fig.canvas.mpl_connect('key_press_event', toggle_color)
        #fig.canvas.start_event_loop(timeout=-5)
        # Display the plot
    
        cid=fig.canvas.mpl_connect('close_event', onclose)
        fig.show() # this call does not block on my system
        fig.canvas.start_event_loop() # block here until window closed
        # After closing the image window, you can access the green and red pixel coordinate lists

        # To select the truck, choose a point on it. Points are input to the model in (x,y) format and come with labels 1 (foreground point) or 0 (background point). Multiple points can be input; here we use only one. The chosen point will be shown as a star on the image.
        #print("Hereeeeeeeee")
        
        # ws['B'+str(c+2)]=str(len(green)) 
        # ws['C'+str(c+2)]=str(len(red))
        # ws['D'+str(c+2)]=str()
        # input_point=np.concatenate((green,red))
        # input_label=np.concatenate(([1]*len(green),[0]*len(red)))
       
        # masks, scores, logits = predictor.predict(
        #     point_coords=input_point,
        #     point_labels=input_label,
        #     multimask_output=True,
        # )
        
        
        
        
    
        
        
        
        
        
        


        #sleep(1)  
        # if np.max(score)<0.8:
        #     print("your score should be more than 0.8, try again")
        #     inc=""
        #     co+=1
        #     if co>=2:
        #         inc=input("you tried more than 10 times\nYou can continue and save the best score ("+str(max(score))+")\nif you want to continue press y")
        # else:
        inc="y"
        print(inc)
        
    
    indx=np.argsort(-np.array(score))
    score=np.array(score)[indx]
    ng=np.array(ng)[indx]
    nr=np.array(nr)[indx]
    stdx=np.array(stdx)[indx]
    stdy=np.array(stdy)[indx]
    for i in range(len(score)):
        coun=1
        for col in ws.iter_cols(min_row=c+2,max_row=c+2, max_col=6+i*5, min_col=2+i*5):
            if coun==1:
                ws[col[0].coordinate]=ng[i]
            elif coun==2:
                ws[col[0].coordinate]=nr[i]
            elif coun==3:
                ws[col[0].coordinate]=stdx[i]
            elif coun==4:
                ws[col[0].coordinate]=stdy[i]
            elif coun==5:
                ws[col[0].coordinate]=score[i]
            coun+=1
    np.save(os.path.join(name,"points",str(c)+"_green"),np.array(gp, dtype=object)[indx])
    np.save(os.path.join(name,"points",str(c)+"_red"),np.array(rp, dtype=object)[indx])
    np.save(os.path.join(name,"masks",str(c)+"_mask"),np.array(msk)[indx])
    np.save(os.path.join(name,"sorts",str(c)+"_sort"),indx)
    c+=1
    contin=input("do u want to continue? press y if you want to continue or anyting otherwise ")
    if not contin =='y':
        wb.save(os.path.join(name,name+'.xlsx')  ) 
        f= True
        file=open(os.path.join(name,"time.txt"),'w')
        file.write(str(float(tim)+(time.time()-t)))
        file.close()
    print("Sample:",c)
wb.save(os.path.join(name,name+'.xlsx')  )  
